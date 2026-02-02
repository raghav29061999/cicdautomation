# src/user_data/__init__.py
"""
user_data

This package supports user-supplied test data (JSON or Excel) as an alternative
to LLM-generated TestData.json.

Flow:
1) parser.py      : bytes + filename -> raw python dict
2) normalizer.py  : raw dict -> TestDataSuite-compatible dict (adds run_id, fixes common drift)
3) validator.py   : validates dict against TestDataSuite (Pydantic)

Design goals:
- Deterministic
- Strict contracts
- Minimal assumptions
"""


-----------

# src/user_data/parser.py
from __future__ import annotations

import io
import json
from typing import Any, Dict, Optional

import openpyxl


class UserTestDataParseError(Exception):
    """Raised when user-provided test data cannot be parsed."""


def parse_user_testdata(filename: str, content: bytes) -> Dict[str, Any]:
    """
    Parse user-provided test data into a raw dict.

    Supported inputs:
    - .json : must be valid JSON object (ideally already TestDataSuite shape)
    - .xlsx : must follow strict sheet layout (see normalizer docs)

    Returns: raw python dict (not yet contract-validated)
    """
    if not filename:
        raise UserTestDataParseError("filename is required")

    if content is None:
        raise UserTestDataParseError("content bytes are required")

    lower = filename.lower().strip()

    if lower.endswith(".json"):
        return _parse_json_bytes(content)

    if lower.endswith(".xlsx"):
        return _parse_xlsx_bytes(content)

    raise UserTestDataParseError(f"Unsupported file type for '{filename}'. Use .json or .xlsx")


def _parse_json_bytes(content: bytes) -> Dict[str, Any]:
    try:
        txt = content.decode("utf-8")
    except Exception as e:
        raise UserTestDataParseError(f"Failed to decode JSON as UTF-8: {e}") from e

    try:
        obj = json.loads(txt)
    except Exception as e:
        raise UserTestDataParseError(f"Invalid JSON: {e}") from e

    if not isinstance(obj, dict):
        raise UserTestDataParseError("Top-level JSON must be an object")
    return obj


def _parse_xlsx_bytes(content: bytes) -> Dict[str, Any]:
    """
    Reads Excel using openpyxl into a raw structured dict.
    This raw dict is then normalized in normalizer.py
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise UserTestDataParseError(f"Failed to read .xlsx: {e}") from e

    # Read sheets if present
    out: Dict[str, Any] = {"_source": "xlsx"}

    out["entities_sheet"] = _read_sheet_as_rows(wb, "entities")
    out["records_sheet"] = _read_sheet_as_rows(wb, "records")
    out["datasets_sheet"] = _read_sheet_as_rows(wb, "datasets")
    out["data_slices_sheet"] = _read_sheet_as_rows(wb, "data_slices")
    out["negative_sheet"] = _read_sheet_as_rows(wb, "negative")

    return out


def _read_sheet_as_rows(wb: openpyxl.Workbook, sheet_name: str) -> list[dict[str, Any]]:
    """
    Reads a worksheet into list[dict] using the first row as headers.
    Missing sheet -> returns empty list (normalizer decides if it’s an error).
    """
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    headers = [h for h in headers if h]  # drop empty headers
    if not headers:
        return []

    out: list[dict[str, Any]] = []
    for r in rows[1:]:
        if r is None:
            continue
        row_obj: dict[str, Any] = {}
        empty = True
        for idx, h in enumerate(headers):
            val = r[idx] if idx < len(r) else None
            if val is not None and str(val).strip() != "":
                empty = False
            row_obj[h] = val
        if not empty:
            out.append(row_obj)

    return out
---
# src/user_data/normalizer.py
from __future__ import annotations

import json
from typing import Any, Dict, List, Tuple


class UserTestDataNormalizeError(Exception):
    """Raised when user-provided data cannot be normalized to TestDataSuite shape."""


def normalize_user_testdata(raw: Dict[str, Any], run_id: str) -> Dict[str, Any]:
    """
    Normalize parsed user data into TestDataSuite-compatible dict.

    If raw appears to already be in TestDataSuite shape, apply light cleanup:
    - enforce run_id
    - normalize linked_test_case -> linked_test_cases
    - ensure selectors is dict
    - ensure data_version default

    If raw is from Excel parse (raw["_source"] == "xlsx"), build TestDataSuite dict
    from the strict worksheet layout.
    """
    if not isinstance(raw, dict):
        raise UserTestDataNormalizeError("raw must be a dict")

    if not run_id:
        raise UserTestDataNormalizeError("run_id is required")

    if raw.get("_source") == "xlsx":
        obj = _normalize_from_xlsx(raw, run_id=run_id)
        return _cleanup_common(obj, run_id=run_id)

    # Assume already JSON-like TestDataSuite object
    obj = dict(raw)
    obj["run_id"] = run_id
    if "data_version" not in obj:
        obj["data_version"] = "TD-v1.0"
    if "generation_intent" not in obj:
        obj["generation_intent"] = {"goal": "provided_test_data", "notes": ["User-supplied data"]}

    return _cleanup_common(obj, run_id=run_id)


# -----------------------------
# Excel normalization
# -----------------------------

def _normalize_from_xlsx(raw: Dict[str, Any], run_id: str) -> Dict[str, Any]:
    """
    Strict Excel layout:

    Sheet: entities
      - entity_name (required)
      - description
      - key_fields (comma-separated)

    Sheet: records
      - entity_name (required)
      - other columns become fields

    Sheet: datasets
      - dataset_id (required)
      - purpose
      - linked_acceptance_criteria (csv)
      - linked_test_cases (csv)
      - entity_refs (csv)
      - data_profile_json (optional JSON string)
      - data_profile_size (optional int) [if json not provided]
      - data_profile_variations (csv) [optional]
      - data_profile_constraints_covered (csv) [optional]

    Sheet: data_slices
      - dataset_id (required)
      - slice_id (required)
      - description
      - selectors_json (JSON string, required-ish)
      - expected_behavior_notes (csv)

    Sheet: negative
      - case_id (required)
      - linked_acceptance_criteria (csv)
      - linked_test_cases (csv)
      - description
      - invalid_records_json (JSON string; list[object] or object)
      - expected_validation_or_handling (csv)

    Output -> TestDataSuite dict
    """
    entities_rows = raw.get("entities_sheet", []) or []
    records_rows = raw.get("records_sheet", []) or []
    datasets_rows = raw.get("datasets_sheet", []) or []
    slices_rows = raw.get("data_slices_sheet", []) or []
    negative_rows = raw.get("negative_sheet", []) or []

    # Build entities map
    entities: List[Dict[str, Any]] = []
    entity_names = set()

    for r in entities_rows:
        name = _str(r.get("entity_name"))
        if not name:
            continue
        entity_names.add(name)
        entities.append({
            "entity_name": name,
            "description": _str(r.get("description")) or "",
            "key_fields": _csv_list(r.get("key_fields")),
            "records": [],  # filled later
        })

    # Attach records by entity_name
    records_by_entity: Dict[str, List[Dict[str, Any]]] = {}
    for rr in records_rows:
        en = _str(rr.get("entity_name"))
        if not en:
            continue
        rec = {k: _primitive(rr.get(k)) for k in rr.keys() if k != "entity_name"}
        # drop empty fields
        rec = {k: v for k, v in rec.items() if v is not None and v != ""}
        records_by_entity.setdefault(en, []).append(rec)

    # If entities sheet missing but records exist, infer entities
    if not entities and records_by_entity:
        for en in sorted(records_by_entity.keys()):
            entities.append({
                "entity_name": en,
                "description": "",
                "key_fields": [],
                "records": [],
            })
            entity_names.add(en)

    # Fill entity records
    for e in entities:
        en = e["entity_name"]
        e["records"] = records_by_entity.get(en, [])

    # Build datasets
    datasets: List[Dict[str, Any]] = []
    for dr in datasets_rows:
        ds_id = _str(dr.get("dataset_id"))
        if not ds_id:
            continue

        data_profile = _parse_optional_json_str(dr.get("data_profile_json"))
        if data_profile is None:
            # Build from optional columns
            data_profile = {
                "size": _int_or_none(dr.get("data_profile_size")),
                "variations": _csv_list(dr.get("data_profile_variations")),
                "constraints_covered": _csv_list(dr.get("data_profile_constraints_covered")),
            }
            # normalize None -> remove
            data_profile = {k: v for k, v in data_profile.items() if v not in (None, [], "")}

        datasets.append({
            "dataset_id": ds_id,
            "purpose": _str(dr.get("purpose")) or "",
            "linked_acceptance_criteria": _csv_list(dr.get("linked_acceptance_criteria")),
            "linked_test_cases": _csv_list(dr.get("linked_test_cases")),
            "entity_refs": _csv_list(dr.get("entity_refs")),
            "data_profile": data_profile if isinstance(data_profile, dict) else {},
            "data_slices": [],  # attach later
        })

    # Attach data_slices by dataset_id
    slices_by_ds: Dict[str, List[Dict[str, Any]]] = {}
    for sr in slices_rows:
        ds_id = _str(sr.get("dataset_id"))
        sl_id = _str(sr.get("slice_id"))
        if not ds_id or not sl_id:
            continue

        selectors = _parse_optional_json_str(sr.get("selectors_json"))
        if selectors is None:
            # selectors is important; keep empty dict but note in expected_behavior_notes
            selectors = {}

        slice_obj = {
            "slice_id": sl_id,
            "description": _str(sr.get("description")) or "",
            "selectors": selectors if isinstance(selectors, dict) else {},
            "expected_behavior_notes": _csv_list(sr.get("expected_behavior_notes")),
        }
        slices_by_ds.setdefault(ds_id, []).append(slice_obj)

    for ds in datasets:
        ds_id = ds["dataset_id"]
        ds["data_slices"] = slices_by_ds.get(ds_id, [])

    # Build negative_and_edge_data
    negative: List[Dict[str, Any]] = []
    for nr in negative_rows:
        cid = _str(nr.get("case_id"))
        if not cid:
            continue

        invalid_records = _parse_optional_json_str(nr.get("invalid_records_json"))
        if invalid_records is None:
            invalid_records = []
        if isinstance(invalid_records, dict):
            invalid_records = [invalid_records]
        if not isinstance(invalid_records, list):
            invalid_records = []

        negative.append({
            "case_id": cid,
            "linked_acceptance_criteria": _csv_list(nr.get("linked_acceptance_criteria")),
            "linked_test_cases": _csv_list(nr.get("linked_test_cases")),
            "description": _str(nr.get("description")) or "",
            "invalid_records": invalid_records,
            "expected_validation_or_handling": _csv_list(nr.get("expected_validation_or_handling")),
        })

    # Construct TestDataSuite dict
    obj: Dict[str, Any] = {
        "run_id": run_id,
        "data_version": "TD-v1.0",
        "generation_intent": {
            "goal": "provided_test_data",
            "notes": ["User-provided test data uploaded via Excel"],
        },
        "entities": entities,
        "datasets": datasets,
        "negative_and_edge_data": negative,
        "assumptions_and_limits": [],
    }

    return obj


# -----------------------------
# Common cleanup (JSON drift)
# -----------------------------

def _cleanup_common(obj: Dict[str, Any], run_id: str) -> Dict[str, Any]:
    """
    Fix common drift while preserving meaning:
    - ensure run_id
    - linked_test_case -> linked_test_cases
    - selectors must be dict
    - ensure list fields are lists
    """
    out = dict(obj)
    out["run_id"] = run_id
    if "data_version" not in out:
        out["data_version"] = "TD-v1.0"
    if "assumptions_and_limits" not in out or out["assumptions_and_limits"] is None:
        out["assumptions_and_limits"] = []

    # normalize datasets
    for ds in out.get("datasets", []) or []:
        if not isinstance(ds, dict):
            continue
        if "linked_test_case" in ds and "linked_test_cases" not in ds:
            ds["linked_test_cases"] = ds.pop("linked_test_case")
        ds["linked_acceptance_criteria"] = _ensure_str_list(ds.get("linked_acceptance_criteria"))
        ds["linked_test_cases"] = _ensure_str_list(ds.get("linked_test_cases"))
        ds["entity_refs"] = _ensure_str_list(ds.get("entity_refs"))

        # slices
        slices = ds.get("data_slices") or []
        if not isinstance(slices, list):
            slices = []
        for sl in slices:
            if not isinstance(sl, dict):
                continue
            selectors = sl.get("selectors")
            if selectors is None:
                sl["selectors"] = {}
            elif not isinstance(selectors, dict):
                # If selectors accidentally came as stringified json, try parse
                if isinstance(selectors, str):
                    parsed = _parse_optional_json_str(selectors)
                    sl["selectors"] = parsed if isinstance(parsed, dict) else {}
                else:
                    sl["selectors"] = {}
            sl["expected_behavior_notes"] = _ensure_str_list(sl.get("expected_behavior_notes"))
        ds["data_slices"] = slices

        # data_profile should be dict
        if not isinstance(ds.get("data_profile"), dict):
            ds["data_profile"] = {}

    # normalize negative/edge
    for neg in out.get("negative_and_edge_data", []) or []:
        if not isinstance(neg, dict):
            continue
        if "linked_test_case" in neg and "linked_test_cases" not in neg:
            neg["linked_test_cases"] = neg.pop("linked_test_case")
        neg["linked_acceptance_criteria"] = _ensure_str_list(neg.get("linked_acceptance_criteria"))
        neg["linked_test_cases"] = _ensure_str_list(neg.get("linked_test_cases"))
        if isinstance(neg.get("invalid_records"), dict):
            neg["invalid_records"] = [neg["invalid_records"]]
        if not isinstance(neg.get("invalid_records"), list):
            neg["invalid_records"] = []

        neg["expected_validation_or_handling"] = _ensure_str_list(neg.get("expected_validation_or_handling"))

    # normalize entities records
    for ent in out.get("entities", []) or []:
        if not isinstance(ent, dict):
            continue
        ent["key_fields"] = _ensure_str_list(ent.get("key_fields"))
        if not isinstance(ent.get("records"), list):
            ent["records"] = []

    return out


# -----------------------------
# Utilities
# -----------------------------

def _str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _csv_list(v: Any) -> List[str]:
    """
    Accepts:
    - None
    - "a,b,c"
    - already list/tuple
    """
    if v is None:
        return []
    if isinstance(v, (list, tuple)):
        return [str(x).strip() for x in v if str(x).strip()]
    s = str(v).strip()
    if not s:
        return []
    # allow semicolon too
    parts = [p.strip() for p in s.replace(";", ",").split(",")]
    return [p for p in parts if p]


def _primitive(v: Any) -> Any:
    # openpyxl returns numbers/bools/dates; keep as-is if JSON-serializable
    if v is None:
        return None
    # convert datetime/date to ISO if present
    try:
        import datetime as _dt
        if isinstance(v, (_dt.datetime, _dt.date)):
            return v.isoformat()
    except Exception:
        pass
    return v


def _int_or_none(v: Any) -> int | None:
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(v)
    except Exception:
        return None


def _parse_optional_json_str(v: Any) -> Any | None:
    if v is None:
        return None
    if isinstance(v, (dict, list)):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:
        return json.loads(s)
    except Exception:
        return None


def _ensure_str_list(v: Any) -> List[str]:
    if v is None:
        return []
    if isinstance(v, list):
        return [str(x).strip() for x in v if str(x).strip()]
    if isinstance(v, tuple):
        return [str(x).strip() for x in v if str(x).strip()]
    s = str(v).strip()
    if not s:
        return []
    return [s]

----
# src/user_data/validator.py
from __future__ import annotations

from typing import Any, Dict

from pydantic import ValidationError

from src.test_data_design.test_data_schema import TestDataSuite


class UserTestDataValidationError(Exception):
    """Raised when user-provided test data fails contract validation."""


def validate_user_testdata(obj: Dict[str, Any]) -> None:
    """
    Validate normalized dict against TestDataSuite contract.
    Raises UserTestDataValidationError with readable error details.
    """
    try:
        TestDataSuite(**obj)
    except ValidationError as e:
        # Keep message readable for UI/CLI
        raise UserTestDataValidationError(_format_pydantic_errors(e)) from e


def _format_pydantic_errors(e: ValidationError) -> str:
    parts = []
    for err in e.errors():
        loc = ".".join(str(x) for x in err.get("loc", []))
        msg = err.get("msg", "")
        typ = err.get("type", "")
        parts.append(f"- {loc}: {msg} ({typ})")
    return "User test data schema validation failed:\n" + "\n".join(parts)
